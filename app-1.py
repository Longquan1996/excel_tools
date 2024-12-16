import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import exceptions as openpyxl_exceptions
from copy import copy

def add_file():
    file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if file and file not in file_list:
        file_list.append(file)
        file_display.insert(tk.END, os.path.basename(file) + "\n")
        update_sheet_options()

def select_folder():
    folder = filedialog.askdirectory()
    if folder:
        for file in os.listdir(folder):
            if file.endswith((".xlsx", ".xls")):
                full_path = os.path.join(folder, file)
                if full_path not in file_list:
                    file_list.append(full_path)
                    file_display.insert(tk.END, os.path.basename(full_path) + "\n")
        update_sheet_options()

def delete_file():
    selected_files = list(file_display.curselection()) # index  list
    if not selected_files:
        messagebox.showwarning("警告", "请选择要删除的文件！")
        return
    for selected in selected_files: # index
        file_list.pop(selected)  # 从 file_list 中移除文件
        file_display.delete(selected)  # 从显示区域中移除
    update_sheet_options()  # 更新页签选择列表

def update_sheet_options():
    if not file_list:
        return
    try:
        first_file = file_list[0]
        wb = load_workbook(first_file, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        sheet_name_list.set(sheet_names)
    except Exception as e:
        messagebox.showerror("错误", f"无法读取页签名称: {e}")

def combine_excel_files():
    if not file_list:
        messagebox.showerror("错误", "请添加至少一个 Excel 文件！")
        return

    selected_sheets = list(sheet_listbox.curselection())
    if not selected_sheets:
        messagebox.showerror("错误", "请选择至少一个页签！")
        return
    
    try:
        combined_df = {}
        title_row_dict = {}

        total = len(selected_sheets) * len(file_list)
        count = 0
        for sheet_idx in selected_sheets:
            sheet_name = sheet_name_list.get()[sheet_idx]
            sheet_data = pd.DataFrame()

            for idx, file_path in enumerate(file_list):
                if idx == 0:  # 第一次读取，确定标题行
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, dtype=str)
                    title_row = None
                    for i, row in df.iterrows():
                        if row.notnull().all():
                            title_row = i + 1
                            break
                    if title_row is None:
                        raise Exception(f"文件 {file_path} 中页签 {sheet_name} 没有找到标题行！")

                    title_row_dict[sheet_name] = title_row
                    df.columns = range(len(df.columns))
                    df = df.iloc[title_row:].reset_index(drop=True)
                    sheet_data = pd.concat([sheet_data, df], ignore_index=True, join='outer')

                else:  # 后续文件跳过标题行
                    df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=title_row_dict[sheet_name]-1, dtype=str)
                    df.columns = range(len(df.columns))
                    sheet_data = pd.concat([sheet_data, df], ignore_index=True, join='outer')
                count += 1
                progress_bar.config(value=count, maximum=total)
                root.update_idletasks() # 更新进度条显示
            combined_df[sheet_name] = sheet_data

        global combined_data, combined_title_rows
        combined_data = combined_df
        combined_title_rows = title_row_dict

        messagebox.showinfo("成功", "合并完成！请另存为文件。")
        save_button.config(state=tk.NORMAL)

    except FileNotFoundError as e:
        messagebox.showerror("文件错误", str(e))
    except ValueError as e:
        messagebox.showerror("数据错误", str(e))
    except openpyxl_exceptions.InvalidFileException:
        messagebox.showerror("文件错误", "检测到无效的 Excel 文件！")
    except Exception as e:
        messagebox.showerror("未知错误", str(e))

def save_combined_file():
    output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if not output_file:
        return

    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for sheet_name, df in combined_data.items():
                title_row = combined_title_rows[sheet_name]
                df.to_excel(writer, sheet_name=sheet_name, header=False, index=False)

                new_wb = writer.book
                new_ws = new_wb[sheet_name]
                new_ws.insert_rows(1, title_row)

                original_wb = load_workbook(file_list[0], data_only=True)
                original_ws = original_wb[sheet_name]

                for row in original_ws.iter_rows(min_row=1, max_row=title_row, values_only=False):
                    for cell in row:
                        new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.fill = copy(cell.fill)
                            new_cell.alignment = copy(cell.alignment)
                            new_cell.border = copy(cell.border)
                            new_cell.number_format = cell.number_format

                for merged_range in original_ws.merged_cells.ranges:
                    min_col, min_row, max_col, max_row = merged_range.bounds
                    if min_row <= title_row:
                        new_ws.merge_cells(start_row=min_row, start_column=min_col,
                                           end_row=max_row, end_column=max_col)

        messagebox.showinfo("成功", f"文件已保存为 {output_file}")

    except Exception as e:
        messagebox.showerror("保存错误", str(e))

# **新窗口：批量转换文件功能**
def batch_convert_files_window():
    # 创建一个新的 Toplevel 窗口
    convert_window = tk.Toplevel(root)
    convert_window.title("待定")
    
    # 设置窗口大小，格式为 "宽度x高度"
    convert_window.geometry("400x300")  # 设置为 400px 宽，300px 高

    label_convert = tk.Label(convert_window, text="待定")
    label_convert.pack(padx=10, pady=10)

    button_convert = tk.Button(convert_window, text="待定")
    button_convert.pack(padx=10, pady=10)

def show_help():
    messagebox.showinfo("帮助", "这是一个功能集合工具！选择一个功能进行操作。")

# 创建主窗口
root = tk.Tk()
root.title("HR表格处理工具集合")

# 创建菜单栏
menu_bar = tk.Menu(root)
root.config(menu=menu_bar)

# 功能菜单
feature_menu = tk.Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="其他功能", menu=feature_menu)
feature_menu.add_command(label="待定", command=batch_convert_files_window)  # 新功能：打开转换窗口
feature_menu.add_separator()
feature_menu.add_command(label="退出", command=root.quit)

# 帮助菜单
# help_menu = tk.Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="使用帮助", command=show_help)
# help_menu.add_command(label="使用帮助", command=show_help)

# 文件选择区域
frame_files = tk.Frame(root)
frame_files.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

file_list = []
combined_data = None
combined_title_rows = None

button_add_folder = tk.Button(frame_files, text="添加文件夹", command=select_folder)
button_add_folder.grid(row=0, column=0, padx=5)

button_add_file = tk.Button(frame_files, text="添加文件", command=add_file)
button_add_file.grid(row=0, column=1, padx=5)

# 使用 Listbox 控件来显示文件名
file_display = tk.Listbox(frame_files, height=10, width=60, font=("Times New Roman", 8), selectmode=tk.SINGLE)
file_display.grid(row=1, column=0, columnspan=2, pady=5)

# 添加删除按钮
button_delete_file = tk.Button(frame_files, text="删除列表文件", command=delete_file)
button_delete_file.grid(row=2, column=0, columnspan=2, pady=5)

# 页签选择区域
frame_sheet = tk.Frame(root)
frame_sheet.pack(pady=10, padx=10)

label_sheet = tk.Label(frame_sheet, text="选择需要合并的页签：")
label_sheet.grid(row=0, column=0, padx=5, pady=5)

sheet_name_list = tk.Variable()
sheet_listbox = tk.Listbox(frame_sheet, listvariable=sheet_name_list, selectmode=tk.MULTIPLE, height=8, width=40, font=("Times New Roman", 8))
sheet_listbox.grid(row=1, column=0, padx=5, pady=5)

# 进度条
progress_bar = ttk.Progressbar(root, length=300, mode="determinate")
progress_bar.pack(pady=10)
# progress_bar_in = ttk.Progressbar(root, mode='indeterminate', length=200)
# progress_bar_in.pack(pady=10)

# 合并和保存按钮
button_combine = tk.Button(root, text="开始合并", command=combine_excel_files, width=20, bg="lightblue")
button_combine.pack(side=tk.LEFT, pady=5)

save_button = tk.Button(root, text="另存为", command=save_combined_file, width=20, bg="lightgreen", state=tk.DISABLED)
save_button.pack(side=tk.RIGHT, pady=5)

root.mainloop()
