import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook

# 用于保存汇总的数据
combined_df = {}

def select_folder():
    folder = filedialog.askdirectory(title="选择Excel文件夹")
    if folder:
        if not any(f.endswith(('.xlsx', '.xls')) for f in os.listdir(folder)):
            messagebox.showwarning("警告", "文件夹中没有 Excel 文件！")
            return
        folder_path.set(folder)
        update_file_list(folder)
    else:
        messagebox.showwarning("警告", "未选择文件夹！")

def update_file_list(folder):
    listbox.delete(0, tk.END) 
    excel_files = [f for f in os.listdir(folder) if f.endswith(('.xlsx', '.xls'))]
    for file in excel_files:
        listbox.insert(tk.END, file)
    file_count_label.set(f"共找到 {len(excel_files)} 个文件")
    if not excel_files:
        messagebox.showinfo("提示", "文件夹中没有 Excel 文件！")

def update_status(message):
    status_label.config(text=message)
    root.update_idletasks()  # 更新界面，显示最新状态

def update_progress_bar(value, total):
    progress_bar.config(value=value, maximum=total)
    root.update_idletasks()  # 更新进度条显示

def get_title_row(df):
    pass

def get_sheet_names():
    """获取所有工作表的名称"""
    folder = folder_path.get()
    for file_name in os.listdir(folder):
        if file_name.endswith(('.xlsx', '.xls')):  # excel
            file_path = os.path.join(folder, file_name)
            try:
                wb = load_workbook(file_path, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
                # for i in sheet_names:
                #     listbox_sheet_names.insert(tk.END, i)
                # listbox_sheet_names.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)
                return sheet_names
            except Exception as e:
                messagebox.showerror("错误", f"错误信息: {e}")
                return []

def combine_sheets():
    folder = folder_path.get()
    sheet = sheet_name.get()

    if not folder:
        messagebox.showerror("错误", "请选择文件夹！")
        return

    count = 0
    total_files = sum(1 for f in os.listdir(folder) if f.endswith(('.xlsx', '.xls')))

    update_status(f"正在汇总页签 '{sheet}' ...")  # 更新状态栏显示汇总过程

    for idx, file_name in enumerate(os.listdir(folder)): # 遍历所有excel
        if file_name.endswith(('.xlsx', '.xls')):  # excel
            file_path = os.path.join(folder, file_name)
            if count == 0: # 第一个表格
                new_wb = Workbook()
                new_ws = new_wb.create_sheet(title=sheet_name)
                try:
                    original_wb = load_workbook(file_path, data_only=True)
                    original_ws = original_wb[sheet_name]
                    # df = pd.read_excel(file_path, sheet_name=sheet, skiprows=skip, dtype={'证件号码': str})
                    sheet_data = pd.concat([sheet_data, df], ignore_index=True)
                    count += 1
                except PermissionError:
                    update_status(f"无法访问文件: {file_name}")
                except ValueError:
                    update_status(f"文件 {file_name} 中没有找到页签: {sheet_name}")
                except (FileNotFoundError, OSError):
                    update_status(f"文件 {file_name} 无法打开或不存在")
                except Exception as e:
                    update_status(f"读取文件 {file_name} 时发生未知错误: {e}")

                # 更新进度条
                update_progress_bar(idx + 1, total_files)

    if count == 0:
        messagebox.showerror("错误", f"没有成功处理任何文件中的页签: {sheet}！")
        return

    combined_df[sheet] = sheet_data  # 将汇总的数据存储到字典中，key 为页签名
    update_status(f"汇总页签 '{sheet}' 完成！处理了 {count} 个文件。")
    update_combined_info()  # 更新已汇总页签信息

def save_combined_file():
    if not combined_df:
        messagebox.showerror("错误", "没有任何汇总数据！")
        return

    output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if not output_file:
        messagebox.showinfo("信息", "您没有选择文件，保存操作已取消。")
        return

    try:
        with pd.ExcelWriter(output_file) as writer:
            for sheet_name, df in combined_df.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        messagebox.showinfo("完成", f"汇总文件已保存为: {output_file}")
        update_status("文件保存完成。")  # 文件保存后，更新状态栏
    except Exception as e:
        messagebox.showerror("错误", f"保存文件时发生错误: {e}")

def clear_combined_data():
    global combined_df  # 确保在函数开始时声明 global
    if not combined_df:
        messagebox.showinfo("信息", "没有数据可清空。")
        return

    answer = messagebox.askyesno("确认", "您确定要清空汇总数据吗？")
    if answer:
        combined_df = {}  # 清空汇总结果
        update_status("汇总结果已清空，您可以重新开始操作。")  # 更新状态栏
        update_combined_info()  # 更新已汇总页签信息

def update_combined_info():
    combined_info = "\n".join(combined_df.keys()) if combined_df else "无"
    combined_info_label.config(text=f"已汇总的页签:\n{combined_info}")

root = tk.Tk()
root.title("Excel 汇总工具")

folder_path = tk.StringVar()
tk.Label(root, text="文件夹:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
tk.Entry(root, textvariable=folder_path, width=40).grid(row=0, column=1, padx=5, pady=5)
tk.Button(root, text="选择", command=select_folder).grid(row=0, column=2, padx=5, pady=5)

tk.Label(root, text="文件列表:").grid(row=1, column=0, padx=5, pady=5, sticky="ne")
listbox = tk.Listbox(root, height=10, width=60)
listbox.grid(row=1, column=1, padx=5, pady=5, columnspan=2)
file_count_label = tk.StringVar(value="共找到 0 个文件")
tk.Label(root, textvariable=file_count_label).grid(row=2, column=1, sticky="w", padx=5)

tk.Label(root, text="页签名:").grid(row=3, column=0, padx=5, pady=5, sticky="e")
sheet_name = tk.StringVar(value="1、初试")
tk.Entry(root, textvariable=sheet_name, width=40).grid(row=3, column=1, padx=5, pady=5)

tk.Label(root, text="标题行:").grid(row=4, column=0, padx=5, pady=5, sticky="e")
title_row = tk.StringVar(value="3")
tk.Entry(root, textvariable=title_row, width=40).grid(row=4, column=1, padx=5, pady=5)

# 状态栏标签
status_label = tk.Label(root, text="欢迎使用 Excel 汇总工具！", relief="sunken", anchor="w", padx=5, pady=5)
status_label.grid(row=7, column=0, columnspan=3, sticky="ew")

# 进度条
progress_bar = ttk.Progressbar(root, length=200, mode="determinate")
progress_bar.grid(row=5, column=0, columnspan=3, pady=10)  # 放到新的一行，并增加适当的间距

# 已汇总的页签信息
combined_info_label = tk.Label(root, text="已汇总的页签:\n无", anchor="w", padx=5, pady=5)
combined_info_label.grid(row=8, column=0, columnspan=3, sticky="ew", padx=5)

tk.Button(root, text="开始汇总", command=combine_sheets).grid(row=6, column=1, columnspan=3, pady=10, sticky="w")  # 放在下一行
tk.Button(root, text="保存文件", command=save_combined_file).grid(row=6, column=0, columnspan=3, pady=10)  # 放在下一行
tk.Button(root, text="清空汇总结果", command=clear_combined_data).grid(row=6, column=2, columnspan=3, pady=10, sticky="w")  # 放在下一行

try:
    root.mainloop()  # Tkinter 主循环
except Exception as e:
    print(f"程序出现错误: {e}")
    input("按 Enter 键退出...")
